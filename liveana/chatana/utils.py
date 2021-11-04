from chat_downloader import ChatDownloader
def chat_record(url, count=None):
    chats_time = {}
    #url = 'https://www.youtube.com/watch?v=ljUsRNYzl0k'
    chat = ChatDownloader().get_chat(url, max_messages=count)

    for message in chat:
        message = chat.format(message)
        message = message.split(' ')
        min = message[0][:-3]
        if min not in chats_time:
            chats_time[min] = 1
        else:
            chats_time[min] += 1

    return chats_time

import requests
from bs4 import BeautifulSoup as bs
def get_title(url):
    web = requests.get(url)
    soup = bs(web.text, 'html.parser')
    title = soup.find('title')
    #print(title.string)
    return title.string

import openpyxl
from openpyxl.chart import LineChart, Reference
def wExcel(data, title='chat-count'):
    wb = openpyxl.Workbook()
    ws = wb.active

    time = [t for t in data]
    count = [data[time] for time in data]

    rows = [[time[i], count[i]] for i in range(len(time))]
    rows.insert(0, ['time', 'count'])
    #print(rows)
    for  row in rows:
        ws.append((row))

    chart = LineChart()
    chart.title = "留言量表"
    chart.x_axis.title = '分鐘'
    chart.y_axis.title = '留言數'
    data = Reference(ws, min_col=2, min_row=1, max_row=len(ws['B']))
    chart.add_data(data, titles_from_data=True)
    time = Reference(ws, min_col=1, min_row=2, max_row=len(ws['A']))
    chart.set_categories(time)

    ws.add_chart(chart, 'D1')
    #wb.save(f'{title}.xlsx')

    return wb

import io
def get_stream(file):
    excel_stream = io.BytesIO()
    file.save(excel_stream)
    res = excel_stream.getvalue()
    excel_stream.close()
    return res


if __name__ == '__main__':
    url='https://www.youtube.com/watch?v=ljUsRNYzl0k'
    data = chat_record(url,100)
    print(data)
    title = get_title(url)
    wExcel(data,title)